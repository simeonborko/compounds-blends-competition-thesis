from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class Table:

    _NAME = None
    _FIELDS = None
    _PRIMARY = None  # pocet stlpcov zo zaciatku

    __REDFILL = PatternFill("solid", fgColor="FF0000")
    __YELLOWFILL = PatternFill("solid", fgColor="FFFF00")

    def __init__(self, wb: Workbook, conn):
        self.__wb = wb
        self.__conn = conn

    def __select_cursor(self):
        c = self.__conn.cursor()
        c.execute("SELECT {} FROM {}".format(','.join(self._FIELDS), self._NAME))
        return c

    def __update(self, datarow):
        query = "UPDATE {} SET {} WHERE {}".format(
            self._NAME,
            ', '.join(f + " = %s" for f in self._FIELDS[self._PRIMARY:]),
            ' AND '.join(f + " = %s" for f in self._FIELDS[:self._PRIMARY])
        )
        data = datarow[self._PRIMARY:] + datarow[:self._PRIMARY]
        c = self.__conn.cursor()
        c.execute(query, data)

    def create_sheet(self):
        if self._NAME in self.__wb.get_sheet_names():
            raise Exception('Harok s nazvom {} uz existuje'.format(self._NAME))
        sheet = self.__wb.create_sheet(self._NAME)

        # nadpisy
        sheet.append(self._FIELDS)
        sheet.row_dimensions[1].font = Font(bold=True)
        for i in range(self._PRIMARY):
            sheet.cell(row=1, column=i+1).font = Font(bold=True, italic=True)
        sheet.iter_rows()

        # data
        for row in self.__select_cursor():
            sheet.append(row)

    def __split_fields(self) -> (tuple, tuple):
        """Rozdeli stlpce podla toho, ci su editovatelne alebo generovane"""
        editable = []
        generated = []
        for i, field in enumerate(self._FIELDS[self._PRIMARY:], self._PRIMARY):
            if field[:2] == 'G_' and field[-8:] != '__ignore':
                generated.append(i)
            else:
                editable.append(i)
        return tuple(editable), tuple(generated)

    def sync(self):

        db = {dbvals[:self._PRIMARY]: list(dbvals) for dbvals in self.__select_cursor()}
        new_db_keys = set(db.keys())
        editable, generated = self.__split_fields()

        sheet = self.__wb[self._NAME]
        for excelrow in sheet.iter_rows(min_row=2):
            excelvals = tuple(cell.value for cell in excelrow)
            key = excelvals[:self._PRIMARY]

            if key not in db:
                print('Harok: {}, riadok: {}. Riadok nenajdeny v databaze.'.format(self._NAME, str(key)))
                for cell in excelrow:
                    cell.fill = self.__REDFILL
                continue

            dbentry = db[key]
            new_db_keys.remove(key)

            # ktore su upravene v tabulke, zmenime aj v databaze
            dbentry_modified = False
            for i in editable:
                if excelvals[i] != dbentry[i]:
                    dbentry[i] = excelvals[i]
                    dbentry_modified = True
            if dbentry_modified:
                self.__update(dbentry)

            # ktore su nanovo vygenerovane v databaze, zmenime v tabulke
            for i in generated:
                if excelvals[i] != dbentry[i]:
                    excelrow[i].value = dbentry[i]
                    excelrow[i].fill = self.__YELLOWFILL

        color_from_row = None
        # ak su nejake nove riadky v db, vypisat
        for key in new_db_keys:
            sheet.append(db[key])
            if color_from_row is None:
                color_from_row = sheet.max_row

        if color_from_row is not None:
            for row in sheet.iter_rows(min_row=color_from_row):
                for cell in row:
                    cell.fill = self.__YELLOWFILL





class ImageTable(Table):
    _NAME = 'image'
    _FIELDS = ('image_id', 'sub_sem_cat', 'dom_sem_cat', 'sub_name', 'dom_name', 'sub_number', 'dom_number', 'half_number', 'sub_sub')
    _PRIMARY = 1


class LanguageTable(Table):
    _NAME = 'language'
    _FIELDS = ('code', 'name')
    _PRIMARY = 1


class NamingUnitTable(Table):
    _NAME = 'naming_unit_test'
    _FIELDS = ('nu_graphic', 'first_language', 'survey_language', 'image_id', 'nu_word_class', 'nu_phonetic', 'nu_syllabic', 'nu_graphic_len', 'nu_phonetic_len', 'nu_syllabic_len', 'n_of_overlapping_letters', 'n_of_overlapping_phones', 'sw1_graphic', 'sw2_graphic', 'sw3_graphic', 'sw4_graphic', 'sw1_headmod', 'sw2_headmod', 'sw3_headmod', 'sw4_headmod', 'sw1_subdom', 'sw2_subdom', 'sw3_subdom', 'sw4_subdom', 'wf_process', 'lexsh_main', 'lexsh_sm', 'lexsh_whatm', 'split_point_1', 'split_point_2', 'split_point_3', 'G_nu_syllabic', 'G_nu_syllabic__ignore', 'G_nu_graphic_len', 'G_nu_phonetic_len', 'G_nu_syllabic_len', 'G_n_of_overlapping_letters', 'G_n_of_overlapping_phones', 'G_lexsh_main', 'G_lexsh_main__ignore', 'G_lexsh_sm', 'G_lexsh_sm__ignore', 'G_lexsh_whatm', 'G_lexsh_whatm__ignore', 'G_split_point_1', 'G_split_point_2', 'G_split_point_3')
    _PRIMARY = 4


class RespondentTable(Table):
    _NAME = 'respondent'
    _FIELDS = ('respondent_id', 'first_language', 'survey_language', 'first_language_original', 'second_language', 'other_language', 'age', 'sex', 'sex_original', 'employment', 'education', 'birth_place', 'year_of_birth', 'responding_date')
    _PRIMARY = 1


class ResponseTable(Table):
    _NAME = 'response'
    _FIELDS = ('respondent_id', 'image_id', 'nu_graphic')
    _PRIMARY = 3


class SourceWordTable(Table):
    _NAME = 'source_word_test'
    _FIELDS = ('sw_graphic', 'first_language', 'survey_language', 'source_language', 'sw_phonetic', 'sw_word_class', 'sw_syllabic', 'sw_graphic_len', 'sw_phonetic_len', 'sw_syllabic_len', 'frequency_in_snc', 'G_sw_syllabic', 'G_sw_syllabic__ignore', 'G_sw_graphic_len', 'G_sw_phonetic_len', 'G_sw_syllabic_len')
    _PRIMARY = 3


class SplinterTable(Table):
    _NAME = 'splinter_test'
    _FIELDS = ('nu_graphic', 'first_language', 'survey_language', 'image_id', 'type_of_splinter', 'sw1_splinter', 'sw2_splinter', 'sw3_splinter', 'sw4_splinter', 'sw1_splinter_len', 'sw2_splinter_len', 'sw3_splinter_len', 'sw4_splinter_len', 'G_sw1_splinter', 'G_sw1_splinter__ignore', 'G_sw2_splinter', 'G_sw2_splinter__ignore', 'G_sw3_splinter', 'G_sw3_splinter__ignore', 'G_sw4_splinter', 'G_sw4_splinter__ignore', 'G_sw1_splinter_len', 'G_sw1_splinter_len__ignore', 'G_sw2_splinter_len', 'G_sw2_splinter_len__ignore', 'G_sw3_splinter_len', 'G_sw3_splinter_len__ignore', 'G_sw4_splinter_len', 'G_sw4_splinter_len__ignore')
    _PRIMARY = 5



# from tools.tools import Connection
#
# with Connection() as conn:
#     c = conn.cursor()
#     c.execute('DESCRIBE image')
#     print(tuple(x[0] for x in list(c) if x[3] == 'PRI'))
