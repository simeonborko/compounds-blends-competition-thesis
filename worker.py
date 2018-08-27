import warnings
from genericpath import isfile
from tkinter import NORMAL, messagebox, DISABLED

from openpyxl import Workbook, load_workbook

import configuration
from syncmanager import SyncManager
from tools import Connection


def __msg(syncmanager: SyncManager, msgs=None) -> str:
    if msgs is None:
        msgs = []

    clss = syncmanager.modified
    if len(clss) > 0:
        msgs.append('Zmenené:')
        msgs.append('\n'.join('- ' + cls.name() for cls in clss))

    clss = syncmanager.stayed
    if len(clss) > 0:
        msgs.append('Bez zmeny:')
        msgs.append('\n'.join('- ' + cls.name() for cls in clss))

    clss = syncmanager.affected
    if len(clss) > 0:
        msgs.append('Zmenené v dôsledku:')
        msgs.append('\n'.join('- ' + cls.name() for cls in clss))

    return '\n\n'.join(msgs)


class Disabler:
    def __init__(self, buttons, checkers):
        self.__buttons = buttons
        self.__checkers = checkers

    def __enter__(self):
        for button in self.__buttons:
            button.configure(state=DISABLED)
        for checker in self.__checkers:
            checker.configure(state=DISABLED)

    def __exit__(self, exc_type, exc_val, exc_tb):
        for button in self.__buttons:
            button.configure(state=NORMAL)
        for checker in self.__checkers:
            checker.configure(state=NORMAL)


def export(clss, widgets):
    with Disabler(**widgets):
        existed = isfile(configuration.XLSX_FILE)

        if not existed:
            wb = Workbook()
            wb.remove(wb['Sheet'])
        else:
            wb = load_workbook(configuration.XLSX_FILE)

        created = []
        notcreated = []
        with Connection() as conn:
            for cls in clss:
                (created if cls(wb, conn).create_sheet() else notcreated).append(cls)

        wb.save(configuration.XLSX_FILE)

        if not existed:
            messagebox.showinfo(
                'Exportovať',
                'Dokument vytvorený:\n\n{}'.format(configuration.XLSX_FILE)
            )
        elif len(created) > 0:
            messagebox.showinfo(
                'Exportovať',
                'Boli pridané tieto hárky:\n\n{}'.format(
                    '\n'.join('- ' + cls.name() for cls in created)
                )
            )
        else:
            messagebox.showwarning(
                'Exportovať',
                'Dokument bol už vytvorený a žiadne hárky neboli pridané'
            )


def sync(clss, unhighlight: bool, widgets):
    with Disabler(**widgets):
        # kontrola ci subor existuje
        if not isfile(configuration.XLSX_FILE):
            messagebox.showerror('Chyba', 'Súbor {} neexistuje'.format(configuration.XLSX_FILE))
            return

        if len(clss) == 0:
            messagebox.showwarning(
                'Synchronizácia',
                'Neboli vybrané žiadne tabuľky'
            )
            return

        wb = load_workbook(configuration.XLSX_FILE)
        with Connection() as conn:
            syncmanager = SyncManager(clss, wb, conn)
            syncmanager.sync(unhighlight)
            conn.commit()

        wb.save(configuration.XLSX_FILE)

        messagebox.showinfo('Synchronizácia', __msg(syncmanager))


def generate(clss, unhighlight: bool, force, corpus, widgets):
    with Disabler(**widgets):
        if not isfile(configuration.XLSX_FILE):
            messagebox.showerror('Chyba', 'Súbor {} neexistuje'.format(configuration.XLSX_FILE))
            return

        if len(clss) == 0:
            messagebox.showwarning(
                'Automatizované vyplnenie',
                'Neboli vybrané žiadne tabuľky'
            )
            return

        wb = load_workbook(configuration.XLSX_FILE)
        with Connection() as conn:
            syncmanager = SyncManager(clss, wb, conn)
            syncmanager.generate(unhighlight, force=force, corpus=corpus)
            conn.commit()

        wb.save(configuration.XLSX_FILE)

        messagebox.showinfo('Automatizované vyplnenie', __msg(syncmanager))


def integrity(clss, unhighlight: bool, widgets):
    with Disabler(**widgets):
        if not isfile(configuration.XLSX_FILE):
            messagebox.showerror('Chyba', 'Súbor {} neexistuje'.format(configuration.XLSX_FILE))
            return

        if len(clss) == 0:
            messagebox.showwarning(
                'Kontrola súdržnosti',
                'Neboli vybrané žiadne tabuľky'
            )
            return

        warnings.filterwarnings('ignore', r'.*Duplicate entry')

        wb = load_workbook(configuration.XLSX_FILE)
        with Connection() as conn:
            syncmanager = SyncManager(clss, wb, conn)
            result = syncmanager.integrity(unhighlight)
            conn.commit()

        warnings.resetwarnings()

        wb.save(configuration.XLSX_FILE)

        msg = __msg(syncmanager, ['\n'.join(
            '- {} (pridané: {}, odobrané: {})'.format(name, *args) for name, args in result.items()
        )])
        messagebox.showinfo('Kontrola súdržnosti', msg)


def syncview(cls, unhighlight: bool, title, widgets):
    with Disabler(**widgets):
        if not isfile(configuration.XLSX_FILE):
            messagebox.showerror('Chyba', 'Súbor {} neexistuje'.format(configuration.XLSX_FILE))
            return

        wb = load_workbook(configuration.XLSX_FILE)

        # ak neexistuje, tak vytvorime
        if cls.name() not in wb.sheetnames:
            with Connection() as conn:
                result = cls(wb, conn).create_sheet()
            if result:
                wb.save(configuration.XLSX_FILE)
                messagebox.showinfo(title, 'Hárok bol vytvorený')
            else:
                messagebox.showerror(title, 'Hárok sa nepodarilo vytvoriť')

        # ak existuje, tak synchronizujeme
        else:
            with Connection() as conn:
                syncmanager = SyncManager([cls], wb, conn)
                syncmanager.sync(unhighlight)
                conn.commit()
            wb.save(configuration.XLSX_FILE)

            messagebox.showinfo(title, __msg(syncmanager))
