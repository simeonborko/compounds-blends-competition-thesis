from unittest import TestCase

from openpyxl import Workbook, load_workbook

import configuration as c

import pymysql


class MyTestCase(TestCase):

    conn = None
    wb = None

    @classmethod
    def setUpClass(cls):
        cls.conn = pymysql.connect(
            host=c.HOST,
            port=c.PORT,
            user='blanksimeon',
            password=c.PASSWD,
            database='blanksimeon'
        )

    @classmethod
    def tearDownClass(cls):
        cls.conn.close()

    def setUp(self):
        self.wb = Workbook()

    def tearDown(self):
        self.conn.rollback()

    def _reopen_wb(self):
        self.wb.save('/tmp/test2.xlsx')
        self.wb = load_workbook('/tmp/test2.xlsx')
