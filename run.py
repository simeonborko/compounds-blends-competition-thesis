from collections import OrderedDict
from tkinter import messagebox
from tkinter import *

from openpyxl import Workbook, load_workbook
from os.path import isfile

from manager import SyncManager
from model import ImageTable, LanguageTable, NamingUnitTable, RespondentTable, ResponseTable, SourceWordTable,\
    SplinterTable, SplinterView, Overview
from tools import Connection, ListSaver
import configuration


TABLES = (ImageTable, LanguageTable, NamingUnitTable, RespondentTable, ResponseTable, SourceWordTable, SplinterTable,
          SplinterView)
SYNC_CLSS = (ImageTable, LanguageTable, NamingUnitTable, RespondentTable, ResponseTable, SourceWordTable, SplinterTable)
GEN_CLSS = (NamingUnitTable, SourceWordTable, SplinterTable)
INTEG_CLSS = (NamingUnitTable, SourceWordTable, SplinterTable)
VIEW_CLSS = (SplinterView, Overview)


def export():

    created = isfile(configuration.XLSX_FILE)

    if not created:
        wb = Workbook()
        wb.remove(wb['Sheet'])
    else:
        wb = load_workbook(configuration.XLSX_FILE)

    with Connection() as conn:
        results = [table.name() for table in TABLES if table(wb, conn).create_sheet()]

    wb.save(configuration.XLSX_FILE)

    if not created:
        messagebox.showinfo(
            'Exportovať',
            'Dokument vytvorený:\n\n{}'.format(configuration.XLSX_FILE)
        )
    elif len(results) > 0:
        messagebox.showinfo(
            'Exportovať',
            'Boli pridané tieto hárky:\n\n{}'.format(
                '\n'.join('- ' + n for n in results)
            )
        )
    else:
        messagebox.showwarning(
            'Exportovať',
            'Dokument bol už vytvorený a žiadne hárky neboli pridané'
        )


def sync():

    # kontrola ci subor existuje
    if not isfile(configuration.XLSX_FILE):
        messagebox.showerror('Chyba', 'Súbor {} neexistuje'.format(configuration.XLSX_FILE))
        return

    # vybrate tabulky
    tables = selected_tables()
    if len(tables) == 0:
        messagebox.showwarning(
            'Synchronizácia',
            'Neboli vybrané žiadne tabuľky'
        )
        return

    dm = DependencyManager()
    synced_tables = []

    wb = load_workbook(configuration.XLSX_FILE)
    with Connection() as conn:

        for table in tables:
            if table(wb, conn).sync():
                synced_tables.append(table)
                dm.set_synced(table)

        for table in dm.effected:
            if table(wb, conn).sync():
                synced_tables.append(table)

        conn.commit()

    wb.save(configuration.XLSX_FILE)

    messagebox.showinfo(
        'Synchronizácia',
        'Synchronizované:\n\n' + '\n'.join('- ' + table.name() for table in synced_tables)
    )


def generate():
    if not isfile(configuration.XLSX_FILE):
        messagebox.showerror('Chyba', 'Súbor {} neexistuje'.format(configuration.XLSX_FILE))
        return

    generable = (NamingUnitTable, SourceWordTable)

    tables = set(selected_tables()) & set(generable)
    if len(tables) == 0:
        messagebox.showwarning(
            'Automatizované vyplnenie',
            'Automatizovane vyplniť sa dajú iba tieto tabuľky:\n\n' + '\n'.join('- ' + table.name() for table in generable)
        )
        return

    force = bool(force_var.get())
    corpus = bool(corpus_var.get())

    affected = OrderedDict()

    dm = DependencyManager()

    wb = load_workbook(configuration.XLSX_FILE)
    with Connection() as conn:
        for table in tables:
            t = table(wb, conn)
            t.sync()
            affected[table.name()] = t.generate(force=force, corpus=corpus)
            if t.sync():
                dm.set_synced(table)

        for table in dm.effected:
            table(wb, conn).sync()

        conn.commit()
    wb.save(configuration.XLSX_FILE)

    messagebox.showinfo(
        'Automatizované vyplnenie',
        'Automatizovane vyplnené a synchronizované:\n\n{}\n\nAj už vyplnené: {}\nKorpus: {}'.format(
            '\n'.join('- {} ({})'.format(tablename, aff) for tablename, aff in affected.items()),
            'Áno' if force else 'Nie',
            'Áno' if corpus else 'Nie'
        )
    )


def integrity():
    if not isfile(configuration.XLSX_FILE):
        messagebox.showerror('Chyba', 'Súbor {} neexistuje'.format(configuration.XLSX_FILE))
        return

    integritable = (NamingUnitTable, SourceWordTable, SplinterTable, LanguageTable)

    # NamingUnit musi ist pred Splinter
    selected = set(selected_tables())
    tables = tuple(tbl for tbl in integritable if tbl in selected)
    if len(tables) == 0:
        messagebox.showwarning(
            'Kontrola súdržnosti',
            'Kontrolovať sa dajú iba tieto tabuľky:\n\n' + '\n'.join('- ' + table.name() for table in integritable)
        )
        return

    affected = OrderedDict()

    dm = DependencyManager()

    wb = load_workbook(configuration.XLSX_FILE)
    with Connection() as conn:
        for table in tables:
            t = table(wb, conn)
            t.sync()
            added = t.integrity_add()
            removed = t.integrity_junk()
            affected[table.name()] = (added, removed)
            if t.sync():
                dm.set_synced(table)

        for table in dm.effected:
            table(wb, conn).sync()

        conn.commit()
    wb.save(configuration.XLSX_FILE)

    messagebox.showinfo(
        'Kontrola súdržnosti',
        'Skontrolované a synchronizované:\n\n{}\n\nOdobrané riadky sú v tabuľke `junk [názov tabuľky]`'.format(
            '\n'.join('- {} (pridané: {}, odobrané: {})'.format(tablename, *a) for tablename, a in affected.items()),
        )
    )


checkvars = []

default_checkvals = [1] * len(TABLES) + [1, 1]


def selected_tables():
    return [TABLES[i] for i, var in enumerate(checkvars) if var.get() == 1]


if not configuration.TKINTER_TRACEBACK:
    Tk.report_callback_exception = lambda obj, exc, val, tb: print(val, file=sys.stderr)


def make_line(parent, row):
    Frame(parent, highlightbackground="#aaaaaa", highlightcolor="#aaaaaa", highlightthickness=1, height=1.5).grid(
        row=row, column=0, sticky=W + E, columnspan=2, padx=5
    )

# framecls = Frame
# Frame = lambda mw: framecls(mw, highlightbackground="gray", highlightcolor="gray", highlightthickness=1, bd=0)


with ListSaver(configuration.CHECKBOX_FILE, default_checkvals) as saver:
    mw = Tk()
    mw.wm_title("Workbook Tlačítka")

    row = 0

    exportFrame = Frame(mw)
    exportFrame.grid(row=row, column=0, columnspan=2, sticky=W+E, pady=10, padx=10)

    row += 1
    make_line(mw, row)
    row += 1

    syncLeftFrame = Frame(mw)
    syncLeftFrame.grid(row=row, column=0, pady=10, padx=10)
    syncRightFrame = Frame(mw)
    syncRightFrame.grid(row=row, column=1, pady=10, padx=10, sticky=W)

    row += 1
    make_line(mw, row)
    row += 1

    generateLeftFrame = Frame(mw)
    generateLeftFrame.grid(row=row, column=0, pady=10, padx=10)
    generateRightFrame = Frame(mw)
    generateRightFrame.grid(row=row, column=1, pady=10, padx=10, sticky=W)

    row += 1
    make_line(mw, row)
    row += 1

    integrityLeftFrame = Frame(mw)
    integrityLeftFrame.grid(row=row, column=0, pady=10, padx=10)
    integrityRightFrame = Frame(mw)
    integrityRightFrame.grid(row=row, column=1, pady=10, padx=10, sticky=W)

    row += 1
    make_line(mw, row)
    row += 1

    viewFrame = Frame(mw)
    viewFrame.grid(row=row, column=0, columnspan=2, sticky=W+E, pady=10, padx=10)

    # export
    Button(exportFrame, text='Exportovať').pack(padx=10, pady=10)

    # sync
    for i, cls in enumerate(SYNC_CLSS):
        Checkbutton(syncLeftFrame, text=cls.name()).grid(row=i, sticky=W, padx=(0, 30), pady=2)
    Button(syncRightFrame, text='Synchronizovať').pack(padx=10, pady=10)

    # generate
    for i, cls in enumerate(GEN_CLSS):
        Checkbutton(generateLeftFrame, text=cls.name()).grid(row=i, sticky=W, padx=(0, 30), pady=2)
    Button(generateRightFrame, text='Vyplniť automatizovane').grid(padx=10, pady=10, row=0, column=0, rowspan=2)
    Checkbutton(generateRightFrame, text='Aj už vyplnené').grid(row=0, column=1, sticky=W)
    Checkbutton(generateRightFrame, text='Korpus').grid(row=1, column=1, sticky=W)

    # integrity
    for i, cls in enumerate(INTEG_CLSS):
        Checkbutton(integrityLeftFrame, text=cls.name()).grid(row=i, sticky=W, padx=(0, 30), pady=2)
    Button(integrityRightFrame, text='Kontrola súdržnosti').pack(padx=10, pady=10)

    # views
    for i, cls in enumerate(VIEW_CLSS):
        Button(viewFrame, text=cls.name().upper()).pack(expand=True, side=LEFT, padx=10, pady=10)



    # Button(
    #     mw,
    #     text='Exportovať',
    # ).grid(row=0, column=0, columnspan=2)
    #
    #
    # exportFrame = Frame(mw, ).pack()
    # xxx = Frame(mw, height=100, width=100, highlightbackground="gray", highlightcolor="gray", highlightthickness=1, bd=0).pack()
    # xxx = Frame(mw, height=100, width=100, highlightbackground="gray", highlightcolor="gray", highlightthickness=1, bd=0).pack()


    #
    # for i in range(7):
    #     Checkbutton(mw, text=i).grid(row=i+1, column=0)
    # Button(
    #     mw,
    #     text='Synchronizovať'
    # ).grid(row=1, column=1, rowspan=7)
    #
    # for i in range(3):
    #     Checkbutton(mw, text=i).grid(row=i+8, column=0)
    # Button(
    #     mw,
    #     text='Vyplniť automatizovane',
    # ).grid(row=1, column=8, rowspan=3)
    # for i in range(2):
    #     Checkbutton(mw, text=i).grid(row=)

    # checkvals = list(saver)
    #
    # for i, tbl in enumerate(TABLES):
    #     var = IntVar(value=checkvals.pop(0))
    #     checkvars.append(var)
    #     Checkbutton(mw, text=tbl.name(), variable=var).grid(row=i, column=0, sticky=W, padx=(0, 30), pady=2)
    #
    # force_var = IntVar(value=checkvals.pop(0))
    # corpus_var = IntVar(value=checkvals.pop(0))
    #
    # Button(
    #     mw,
    #     text='Exportovať',
    #     command=export,
    # ).grid(row=0, column=1, padx=5, pady=5, rowspan=2, sticky=E+W)
    #
    # Button(
    #     mw,
    #     text='Synchronizovať',
    #     command=sync,
    # ).grid(row=2, column=1, padx=5, pady=5, rowspan=2, sticky=E+W)
    #
    # Button(
    #     mw,
    #     text='Vyplniť automatizovane',
    #     command=generate
    # ).grid(row=4, column=1, padx=5, pady=5, rowspan=2, sticky=E+W)
    # Checkbutton(mw, text='Aj už vyplnené', variable=force_var).grid(row=4, column=2)
    # Checkbutton(mw, text='Korpus', variable=corpus_var).grid(row=5, column=2)
    #
    # Button(
    #     mw,
    #     text='Kontrola súdržnosti',
    #     command=integrity
    # ).grid(row=6, column=1, padx=5, pady=5, rowspan=2, sticky=E+W)
    #
    mw.mainloop()
    #
    # for i, var in enumerate(checkvars + [force_var, corpus_var]):
    #     saver[i] = var.get()
