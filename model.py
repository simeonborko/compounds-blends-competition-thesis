from abc import ABC, ABCMeta, abstractmethod
import sys
from collections import namedtuple
from contextlib import contextmanager, ExitStack
from itertools import groupby
from typing import Iterable, Tuple, List, Set, Dict, Any

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet import Worksheet
from pymysql.cursors import DictCursor
import configuration
from entity import SourceWord, NamingUnit, Splinter
from splinter_view_field_manager import SplinterViewFieldManager
from tools import sk, en
from tools.exception import ResponseDuplicatesException, ResponseTypeError
import temp_table


class TableLike(ABC):

    _NAME = None
    _FIELDS = None
    _PRIMARY = None  # pocet stlpcov zo zaciatku
    _EXPORT_SELECT = None

    _REDFILL = PatternFill("solid", fgColor="FFFF7575")
    _YELLOWFILL = PatternFill("solid", fgColor="FFFFFC75")
    _NOFILL = PatternFill()

    ExecuteResult = namedtuple('ExecuteResult', ['cursor', 'result'])

    def __init__(self, wb: Workbook, conn, as_affected: bool = False):
        self._wb = wb
        self.__conn = self._conn = conn
        self._as_affected = as_affected

    def _add_header(self, sheet):
        sheet.append(self._FIELDS)
        emphasized = set(self._emphasized_columns)
        for i in range(len(self._FIELDS)):
            sheet.cell(row=1, column=i + 1).font = Font(italic=True) if i in emphasized else Font(bold=True)

    @property
    def _emphasized_columns(self) -> Iterable[int]:
        """Vrati indexy zvyraznenych stlpcov (ktore nemaju byt upravovane)."""
        return range(self._PRIMARY)

    def _execute(self, *args, **kwargs) -> ExecuteResult:
        c = self.__conn.cursor(**kwargs)
        print(args)
        res = c.execute(*args) or 0
        return self.ExecuteResult(c, res)

    def _executemany(self, *args, **kwargs) -> ExecuteResult:
        c = self.__conn.cursor(**kwargs)
        res = c.executemany(*args) or 0
        return self.ExecuteResult(c, res)

    @classmethod
    def _update_in_sheet(cls, db_values: list, sheet_cells: List[Cell], indices: Iterable[int]) -> bool:
        """
        Aktualizuje hodnoty v harku a zafarbi ich na zlto.
        :param db_values: hodnoty v DB
        :param sheet_cells: bunky v harku
        :param indices: indexy stlpcov, pri ktorych ma prioritu DB
        :return: aktualizovalo sa nieco v harku?
        """
        modified = False
        for i in indices:
            db_val = db_values[i]
            sh_val = sheet_cells[i].value
            if db_val != sh_val and (db_val or sh_val):
                # print(sh_val, type(sh_val), db_val, type(db_val), sep='\t')
                sheet_cells[i].value = db_val
                sheet_cells[i].fill = cls._YELLOWFILL
                modified = True
        return modified

    @staticmethod
    def _update_in_db(db_values: list, sheet_cells: List[Cell], indices: Iterable[int]) -> List[int]:
        """
        Zisti, ci je potrebne aktualizovat databazu.
        Vedlajsi efekt, ze meni hodnoty v db_values.
        :param db_values: hodnoty v databaze
        :param sheet_cells: bunky v harku
        :param indices: indexy, ktore hodnoty sa maju prenasat do DB (editable)
        :return: zoznam indexov stlpcov, kde treba aktualizovat hodnotu v DB
        """
        print('Entering _update_in_db')
        mod_idx = []
        for i in indices:
            if db_values[i] != sheet_cells[i].value and (db_values[i] or sheet_cells[i].value):
                db_values[i] = sheet_cells[i].value
                mod_idx.append(i)
        print('Leaving _update_in_db')
        return mod_idx

    @abstractmethod
    def _sync(self, vals_gen) -> bool:
        """
        Synchronizuje dane riadky.
        :param vals_gen: generator na dvojice (zoznam hodnot v DB, zoznam Cell v SHEETE)
        """
        pass

    @staticmethod
    def _same_fill(cell: Cell, fill: PatternFill) -> bool:
        return cell.fill == fill or cell.fill.fgColor == fill.fgColor

    def create_sheet(self) -> bool:
        if self.sheet_created:
            return False

        sheet = self._wb.create_sheet(self._NAME)

        # nadpisy
        self._add_header(sheet)

        # data
        for row in self._execute(self._EXPORT_SELECT).cursor:
            sheet.append(row)

        return True

    def sync(self, unhighlight: bool = False) -> bool:
        """
        Synchronizuje TableLike objekt medzi DB a Workbookom.
        :param unhighlight: ma sa zrusit zlte zvyraznenie vo Workbooku?
        :return: bol Workbook zmeneny?
        """

        # ziskat zaznamy z DB
        db_dict = {dbvals[:self._PRIMARY]: list(dbvals) for dbvals in self._execute(self._EXPORT_SELECT).cursor}

        # ziskat zaznamy zo SHEETU
        sheet = self._wb[self._NAME]
        sheet_dict = {
            tuple(cell.value for cell in rows[:self._PRIMARY]): rows
            for rows in sheet.iter_rows(min_row=2)
        }

        # porovnat mnoziny klucov
        keys_db_only = set(db_dict.keys()) - set(sheet_dict.keys())
        keys_sheet_only = set(sheet_dict.keys()) - set(db_dict.keys())
        keys_both = set(db_dict.keys()) & set(sheet_dict.keys())

        if unhighlight:
            # odstranit predosle zvyraznenie zmien
            for row in sheet_dict.values():
                for cell in row:
                    if self._same_fill(cell, self._YELLOWFILL):
                        cell.fill = self._NOFILL

        # odstranit zvyraznenie cervenych riadkov, ktore su uz zrazu v DB
        for key, row in sheet_dict.items():
            if key in keys_both:
                for cell in row:
                    if self._same_fill(cell, self._REDFILL):
                        cell.fill = self._NOFILL

        modified = False

        # co su iba v DB, pridat zlto
        if len(keys_db_only) > 0:
            modified = True
            starting_row = sheet.max_row + 1
            for key in keys_db_only:
                sheet.append(db_dict[key])
            for row in sheet.iter_rows(min_row=starting_row):
                for cell in row:
                    cell.fill = self._YELLOWFILL

        # co su iba v SHEET, ofarbit cerveno
        if len(keys_sheet_only) > 0:
            for key in keys_sheet_only:
                for cell in sheet_dict[key]:
                    if not self._same_fill(cell, self._REDFILL):
                        cell.fill = self._REDFILL
                        modified = True

        # co su aj aj, synchronizovat, zafarbit zlto zmenene
        return self._sync((db_dict[k], sheet_dict[k]) for k in keys_both) or modified

    @property
    def primary_fields(self) -> Tuple[str, ...]: return self._FIELDS[:self._PRIMARY]

    @classmethod
    def name(cls) -> str: return cls._NAME

    @property
    def fields(self) -> tuple: return self._FIELDS

    @property
    def sheet_created(self) -> bool: return self._NAME in self._wb.sheetnames


class StaticView(TableLike):

    @property
    def _emphasized_columns(self) -> Iterable[int]:
        return range(len(self._FIELDS))

    def _sync(self, vals_gen):
        """Nahradi hodnoty v SHEETE hodnotami z DB"""
        modified = False
        indices = range(len(self._FIELDS))
        for db_values, sheet_cells in vals_gen:
            if self._update_in_sheet(db_values, sheet_cells, indices):
                modified = True
        return modified


class Overview(StaticView):

    _NAME = 'overview'
    _FIELDS = (
        'respondent_id',
        'image_id',
        'nu_original',
        'nu_modified',
        'nu_graphic',
        'wf_process',
        'wfp_specification',
        'age',
        'sex',
        'first_language',
        'survey_language',
        'sub_sem_cat',
        'dom_sem_cat',
        'sub_name',
        'dom_name',
        'sub_number',
        'dom_number',
        'half_number',
        'nu_phonetic',
        'nu_syllabic',
        'nu_graphic_len',
        'nu_phonetic_len',
        'nu_syllabic_len',
        'sw1_graphic',
        'sw2_graphic',
        'sw3_graphic',
        'sw4_graphic',
        'sw1_word_class',
        'sw2_word_class',
        'sw3_word_class',
        'sw4_word_class',
        'sw1_source_language',
        'sw2_source_language',
        'sw3_source_language',
        'sw4_source_language',
        'sw1_proper_name',
        'sw2_proper_name',
        'sw3_proper_name',
        'sw4_proper_name',
        'sw1_graphic_len',
        'sw2_graphic_len',
        'sw3_graphic_len',
        'sw4_graphic_len',
        'sw1_phonetic',
        'sw2_phonetic',
        'sw3_phonetic',
        'sw4_phonetic',
        'sw1_phonetic_len',
        'sw2_phonetic_len',
        'sw3_phonetic_len',
        'sw4_phonetic_len',
        'sw1_syllabic',
        'sw2_syllabic',
        'sw3_syllabic',
        'sw4_syllabic',
        'sw1_syllabic_len',
        'sw2_syllabic_len',
        'sw3_syllabic_len',
        'sw4_syllabic_len',
        'sw1_frequency_in_snc',
        'sw2_frequency_in_snc',
        'sw3_frequency_in_snc',
        'sw4_frequency_in_snc',
        'gs_name',
        'gs_sw1_splinter',
        'gs_sw2_splinter',
        'gs_sw3_splinter',
        'gs_sw4_splinter',
        'gs_sw1_splinter_len',
        'gs_sw2_splinter_len',
        'gs_sw3_splinter_len',
        'gs_sw4_splinter_len',
        'gs_sw1_splinter_len_to_sw_len',
        'gs_sw2_splinter_len_to_sw_len',
        'gs_sw3_splinter_len_to_sw_len',
        'gs_sw4_splinter_len_to_sw_len',
        'gs_sw1_splinter_len_to_nu_len',
        'gs_sw2_splinter_len_to_nu_len',
        'gs_sw3_splinter_len_to_nu_len',
        'gs_sw4_splinter_len_to_nu_len',
        'gm_name',
        'gm_sw1_splinter',
        'gm_sw2_splinter',
        'gm_sw3_splinter',
        'gm_sw4_splinter',
        'gm_sw1_splinter_len',
        'gm_sw2_splinter_len',
        'gm_sw3_splinter_len',
        'gm_sw4_splinter_len',
        'gm_sw1_splinter_len_to_sw_len',
        'gm_sw2_splinter_len_to_sw_len',
        'gm_sw3_splinter_len_to_sw_len',
        'gm_sw4_splinter_len_to_sw_len',
        'gm_sw1_splinter_len_to_nu_len',
        'gm_sw2_splinter_len_to_nu_len',
        'gm_sw3_splinter_len_to_nu_len',
        'gm_sw4_splinter_len_to_nu_len',
        'ps_name',
        'ps_sw1_splinter',
        'ps_sw2_splinter',
        'ps_sw3_splinter',
        'ps_sw4_splinter',
        'ps_sw1_splinter_len',
        'ps_sw2_splinter_len',
        'ps_sw3_splinter_len',
        'ps_sw4_splinter_len',
        'ps_sw1_splinter_len_to_sw_len',
        'ps_sw2_splinter_len_to_sw_len',
        'ps_sw3_splinter_len_to_sw_len',
        'ps_sw4_splinter_len_to_sw_len',
        'ps_sw1_splinter_len_to_nu_len',
        'ps_sw2_splinter_len_to_nu_len',
        'ps_sw3_splinter_len_to_nu_len',
        'ps_sw4_splinter_len_to_nu_len',
        'pm_name',
        'pm_sw1_splinter',
        'pm_sw2_splinter',
        'pm_sw3_splinter',
        'pm_sw4_splinter',
        'pm_sw1_splinter_len',
        'pm_sw2_splinter_len',
        'pm_sw3_splinter_len',
        'pm_sw4_splinter_len',
        'pm_sw1_splinter_len_to_sw_len',
        'pm_sw2_splinter_len_to_sw_len',
        'pm_sw3_splinter_len_to_sw_len',
        'pm_sw4_splinter_len_to_sw_len',
        'pm_sw1_splinter_len_to_nu_len',
        'pm_sw2_splinter_len_to_nu_len',
        'pm_sw3_splinter_len_to_nu_len',
        'pm_sw4_splinter_len_to_nu_len',
    )

    _EXPORT_SELECT = """SELECT {} FROM (
  SELECT
    RTS.respondent_id as respondent_id,
    RTS.age           as age,
    RTS.sex           as sex,
    RTS.nu_original,
    RTS.nu_modified,
    N.*
  FROM
    nu_full N
    LEFT JOIN respondent_response RTS
      ON N.nu_graphic = RTS.nu_graphic
      AND N.image_id = RTS.image_id
      AND N.first_language = RTS.first_language
      AND N.survey_language = RTS.survey_language
) T""".format(', '.join(_FIELDS))

    _PRIMARY = 5

    def __init__(self, wb: Workbook, conn, as_affected: bool = False):
        super().__init__(wb, conn, as_affected)
        temp_table.create_response_combined(conn)
        temp_table.create_respondent_response(conn)


class EditableTableLike(TableLike):

    _EXCLUDE_EDITABLE = None
    _INCLUDE_EDITABLE = None

    def __init__(self, wb: Workbook, conn, as_affected: bool=False):
        super().__init__(wb, conn, as_affected)
        self.__editable, self.__generated = self.__split_fields()

    @staticmethod
    def __looks_like_generated(field) -> bool:
        return field[:2] == 'G_' and field[-8:] != '__ignore'

    def __is_generated(self, field):
        if self._EXCLUDE_EDITABLE and field in self._EXCLUDE_EDITABLE:
            return True
        elif self._INCLUDE_EDITABLE and field in self._INCLUDE_EDITABLE:
            return False
        else:
            return self.__looks_like_generated(field)

    def __split_fields(self) -> (Tuple[int, ...], Tuple[int, ...]):
        editable = []
        generated = []
        for i, field in enumerate(self._FIELDS[self._PRIMARY:], self._PRIMARY):
            (generated if self.__is_generated(field) else editable).append(i)
        return tuple(editable), tuple(generated)

    @property
    def _emphasized_columns(self) -> Iterable[int]:
        return set(super()._emphasized_columns) | set(self.__generated)

    def _sync(self, vals_gen):
        """
        :param vals_gen: generator (zoznam hodnot v DB riadku, zoznam buniek vo Workbook riadku)
        :return:
        """

        print('Entering _sync', self._as_affected)

        modified = False

        # ako v dosledku, ze nieco ine bolo zmenene, takze vsetko treba zmenit v harku
        if self._as_affected:
            indices = range(len(self._FIELDS))
            for db_values, sheet_cells in vals_gen:
                if self._update_in_sheet(db_values, sheet_cells, indices):
                    modified = True

        else:
            for db_values, sheet_cells in vals_gen:
                indexes_to_update = self._update_in_db(db_values, sheet_cells, self.__editable)
                print(indexes_to_update)
                if len(indexes_to_update) > 0:
                    all_data = dict(zip(self.fields, db_values))
                    fields_to_update = [self.fields[idx] for idx in indexes_to_update]
                    self._update(all_data, fields_to_update)
                    modified = True
                if self._update_in_sheet(db_values, sheet_cells, self.__generated):
                    modified = True

        print('Leaving _sync')

        return modified

    @abstractmethod
    def _update(self, data: Dict[str, Any], fields_to_modify: List[str]):
        """
        Aktualizuje data v DB podla `data`.
        :param data: nazov stlpca -> hodnota. Su tu vsetky stlpce.
        :param fields_to_modify: zoznam stlpcov, ktore sa maju zmenit v DB
        """
        pass

    @property
    def generated_fields(self) -> tuple:
        return tuple(self._FIELDS[i] for i in self.__generated)

    @property
    def editable_fields(self) -> tuple:
        return tuple(self._FIELDS[i] for i in self.__editable)


class Table(EditableTableLike, metaclass=ABCMeta):

    # select iba na tie riadky, ktore potrebuju byt generovane
    _GENERATE_SELECT_NEEDED = None

    # select na vsetky riadky pri generovani (napr. Splinter potrebuje aj udaje zo SourceWord)
    _GENERATE_SELECT_ALL = None

    # select na riadky, ake primarne kluce maju byt v danej tabulke
    _INTEGRITY_SELECT = None

    # zoznam stlpcov, pouzitych vo funkcii integrity_junk
    _INTEGRITY_JUNK_FIELDS = None

    def __init__(self, wb: Workbook, conn, as_affected: bool = False):
        super().__init__(wb, conn, as_affected)
        self._EXPORT_SELECT = "SELECT {} FROM {}".format(','.join(self._FIELDS), self._NAME)
        if self._GENERATE_SELECT_ALL is None:  # SplinterTable ma _GENERATE_SELECT_ALL v triede, nie v inite
            self._GENERATE_SELECT_ALL = "SELECT * FROM {}".format(self._NAME)  # treba *, nestaci self._FIELDS
        self._INTEGRITY_JUNK_FIELDS = self._FIELDS

    def _update(self, data: Dict[str, Any], fields_to_modify: List[str]):
        """`data` obsahuju hodnoty, ktore su primarnym klucom a tie, ktore chceme zmenit.
        Tie, ktore sa nemenia, tam nemaju byt."""

        to_set = ["{0} = %({0})s".format(f) for f in fields_to_modify if f in self.editable_fields]
        if len(to_set) == 0:
            print('Warning: No data to set', data)
            return

        query = "UPDATE {} SET {} WHERE {}".format(
            self._NAME,
            ', '.join(to_set),
            ' AND '.join("{0} = %({0})s".format(f) for f in self.primary_fields),
        )
        self._execute(query, data)

    def _generate(self, force: bool, cls) -> int:

        c = self._execute(
            self._GENERATE_SELECT_ALL if force or not self._GENERATE_SELECT_NEEDED else self._GENERATE_SELECT_NEEDED,
            cursor=DictCursor
        ).cursor

        args = []

        for data in c:
            entity = cls(self, data)
            entity.generate()
            if entity.modified:
                args.append(entity.data)

        if len(args):

            query = "UPDATE {} SET {} WHERE {}".format(
                self._NAME,
                ", ".join("{0}=%({0})s".format(g) for g in self.generated_fields),
                " AND ".join("{0}=%({0})s".format(p) for p in self.primary_fields)
            )

            affected = self._executemany(query, args).result

            if len(args) != affected:
                print("Pocet args: {}, pocet affected: {}".format(len(args), affected), file=sys.stderr)
                affected = affected or 0

            return affected

        else:
            return 0

    def integrity_before(self):
        pass

    def integrity_add(self) -> int:
        return self._execute(
            "INSERT IGNORE INTO {} ({}) {}".format(
                self._NAME,
                ', '.join(self.primary_fields),
                self._INTEGRITY_SELECT
            )
        ).result

    def integrity_junk(self) -> int:
        exres = self._execute(
            "SELECT {} FROM {} TBL LEFT JOIN ({}) TMP ON {} WHERE {}".format(
                ', '.join('TBL.' + field for field in self._INTEGRITY_JUNK_FIELDS),
                self._NAME,
                self._INTEGRITY_SELECT,
                ' AND '.join('TBL.{0} = TMP.{0}'.format(p) for p in self.primary_fields),
                ' AND '.join('TMP.{} IS NULL'.format(p) for p in self.primary_fields)
            )
        )
        if not exres.result:
            return 0

        # vytvorit harok
        sheetname = 'junk {}'.format(self._NAME)
        if sheetname not in self._wb.sheetnames:
            sheet = self._wb.create_sheet(sheetname)
            self._add_header(sheet)
        else:
            sheet = self._wb[sheetname]

        # pridat riadky a ulozit kluce
        keys = []
        for row in exres.cursor:
            keys.append(row[:self._PRIMARY])
            sheet.append(row)

        exres = self._executemany(
            """DELETE FROM {} WHERE {}""".format(
                self._NAME,
                ' AND '.join('{} = %s'.format(p) for p in self.primary_fields)
            ),
            keys
        )
        return exres.result


class ImageTable(Table):
    _NAME = 'image'
    _FIELDS = (
        'image_id', 'sub_sem_cat', 'dom_sem_cat', 'sub_name', 'dom_name',
        'sub_number', 'dom_number', 'half_number', 'sub_sub'
    )
    _PRIMARY = 1


class LanguageTable(Table):
    _NAME = 'language'
    _FIELDS = ('code', 'name')
    _PRIMARY = 1
    _INTEGRITY_SELECT = "SELECT first_language code FROM respondent UNION SELECT survey_language code FROM respondent"


class NamingUnitTable(Table):

    __FROM_SW = (
        'sw1_source_language', 'sw2_source_language', 'sw3_source_language', 'sw4_source_language',
        'sw1_word_class', 'sw2_word_class', 'sw3_word_class', 'sw4_word_class',
    )

    _NAME = 'naming_unit'

    _EXCLUDE_EDITABLE = __FROM_SW

    _FIELDS = (
        'nu_graphic', 'first_language', 'survey_language', 'image_id',
        'wf_process', 'wfp_specification',

        'sw1_graphic', 'sw2_graphic', 'sw3_graphic', 'sw4_graphic',
        *__FROM_SW,
        'sw1_headmod', 'sw2_headmod', 'sw3_headmod', 'sw4_headmod',
        'sw1_subdom', 'sw2_subdom', 'sw3_subdom', 'sw4_subdom',

        'nu_word_class', 'nu_phonetic',
        'nu_syllabic', 'G_nu_syllabic', 'G_nu_syllabic__ignore',
        'nu_graphic_len', 'G_nu_graphic_len',
        'nu_phonetic_len', 'G_nu_phonetic_len',
        'nu_syllabic_len', 'G_nu_syllabic_len',
        'n_of_overlapping_letters', 'G_n_of_overlapping_letters',
        'n_of_overlapping_phones', 'G_n_of_overlapping_phones',
        'lexsh_main', 'G_lexsh_main', 'G_lexsh_main__ignore', 'lexsh_sm', 'G_lexsh_sm', 'G_lexsh_sm__ignore',
        'lexsh_whatm', 'G_lexsh_whatm', 'G_lexsh_whatm__ignore',
        'split_point_1', 'G_split_point_1', 'split_point_2', 'G_split_point_2', 'split_point_3', 'G_split_point_3'
    )
    _PRIMARY = 4

    _GENERATE_SELECT_NEEDED = """select * from {} where
  survey_language='SK' and G_nu_syllabic is null
  or G_nu_graphic_len is null
  or nu_phonetic is not null and G_nu_phonetic_len is null
  or survey_language='SK' and G_nu_syllabic_len is null
  or survey_language='EN' and nu_phonetic is not null and G_nu_syllabic_len is null""".format(_NAME)

    _INTEGRITY_SELECT = "SELECT DISTINCT {} FROM respondent_response".format(', '.join(_FIELDS[:_PRIMARY]))

    def __init__(self, wb: Workbook, conn, as_affected: bool = False):
        super().__init__(wb, conn, as_affected)
        self._EXPORT_SELECT = "SELECT {} FROM ({}) T".format(
            ','.join(self._FIELDS),
            """SELECT
    NU.*,
    `SW1`.`sw_word_class`   AS `sw1_word_class`,
    `SW2`.`sw_word_class`   AS `sw2_word_class`,
    `SW3`.`sw_word_class`   AS `sw3_word_class`,
    `SW4`.`sw_word_class`   AS `sw4_word_class`,
    `SW1`.`source_language` AS `sw1_source_language`,
    `SW2`.`source_language` AS `sw2_source_language`,
    `SW3`.`source_language` AS `sw3_source_language`,
    `SW4`.`source_language` AS `sw4_source_language`

  FROM naming_unit NU

    left join `source_word` `SW1`
      on `NU`.`first_language` = `SW1`.`first_language`
         and `NU`.`survey_language` = `SW1`.`survey_language`
         and `NU`.`sw1_graphic` = `SW1`.`sw_graphic`
    left join `source_word` `SW2`
      on `NU`.`first_language` = `SW2`.`first_language`
         and `NU`.`survey_language` = `SW2`.`survey_language`
         and `NU`.`sw2_graphic` = `SW2`.`sw_graphic`
    left join `source_word` `SW3`
      on `NU`.`first_language` = `SW3`.`first_language`
         and `NU`.`survey_language` = `SW3`.`survey_language`
         and `NU`.`sw3_graphic` = `SW3`.`sw_graphic`
    left join `source_word` `SW4`
      on `NU`.`first_language` = `SW4`.`first_language`
         and `NU`.`survey_language` = `SW4`.`survey_language`
         and `NU`.`sw4_graphic` = `SW4`.`sw_graphic`"""
        )
        self._INTEGRITY_JUNK_FIELDS = tuple(f for f in self._FIELDS if f not in self.__FROM_SW)

    # noinspection PyUnusedLocal
    def generate(self, force, **kwargs) -> int:
        return self._generate(force, NamingUnit)

    def integrity_before(self):
        temp_table.create_response_combined(self._conn)
        temp_table.create_respondent_response(self._conn)


class RespondentTable(Table):
    _NAME = 'respondent'
    _FIELDS = (
        'respondent_id', 'first_language', 'survey_language', 'first_language_original', 'second_language',
        'other_language', 'age', 'sex', 'sex_original', 'employment', 'education', 'birth_place', 'year_of_birth',
        'responding_date'
    )
    _PRIMARY = 1


class SourceWordTable(Table):
    _NAME = 'source_word'
    _FIELDS = ('sw_graphic', 'first_language', 'survey_language',
               'source_language', 'proper_name', 'sw_phonetic', 'G_sw_phonetic', 'sw_word_class',
               'sw_syllabic', 'G_sw_syllabic', 'G_sw_syllabic__ignore',
               'sw_graphic_len', 'G_sw_graphic_len',
               'sw_phonetic_len', 'G_sw_phonetic_len',
               'sw_syllabic_len', 'G_sw_syllabic_len', 'frequency_in_snc')
    _PRIMARY = 3
    _EXCLUDE_EDITABLE = {'frequency_in_snc'}
    _GENERATE_SELECT_NEEDED = """select * from {} where
  survey_language='SK' and G_sw_syllabic is null
  or G_sw_graphic_len is null
  or sw_phonetic is not null and G_sw_phonetic_len is null
  or survey_language='SK' and G_sw_syllabic_len is null
  or survey_language='EN' and sw_phonetic is not null and G_sw_syllabic_len is null
  or survey_language='SK' and frequency_in_snc is null
  or survey_language='EN' and G_sw_phonetic is null""".format(_NAME)

    _INTEGRITY_SELECT = ' UNION '.join(
        'SELECT {0} sw_graphic, first_language, survey_language FROM naming_unit WHERE {0} IS NOT NULL'.format(
            'sw{}_graphic'.format(i+1)
        ) for i in range(4)
    )
    # SELECT sw1_graphic sw_graphic, first_language, survey_language from naming_unit WHERE sw1_graphic is not null
    # UNION SELECT sw2_graphic, first_language, survey_language from naming_unit WHERE sw2_graphic is not null
    # UNION SELECT sw3_graphic, first_language, survey_language from naming_unit WHERE sw3_graphic is not null
    # UNION SELECT sw4_graphic, first_language, survey_language from naming_unit WHERE sw4_graphic is not null

    @staticmethod
    @contextmanager
    def _sk_corpus_context_manager():
        with sk.Corpus(configuration.CORPUS_FILE) as corpus:
            SourceWord.CORPUS = corpus
            try:
                yield
            finally:
                SourceWord.CORPUS = None

    @staticmethod
    @contextmanager
    def _en_corpus_context_manager():
        with en.BncCorpus(configuration.BNC_CORPUS_FILE, configuration.BNC_CORPUS_NOT_FOUND_LIST_FILE) as bnc_corpus:
            SourceWord.BNC_CORPUS = bnc_corpus
            try:
                yield
            finally:
                SourceWord.BNC_CORPUS = None

    @staticmethod
    @contextmanager
    def _en_transcription_context_manager():
        with en.TranscriptionManager() as trans_man:
            SourceWord.TRANSCRIPTION_MANAGER = trans_man
            try:
                yield
            finally:
                SourceWord.TRANSCRIPTION_MANAGER = None

    def generate(self, force, **kwargs) -> int:

        with ExitStack() as stack:

            if kwargs['corpus']:
                stack.enter_context(self._sk_corpus_context_manager())
            if kwargs['bnc_corpus']:
                stack.enter_context(self._en_corpus_context_manager())
            if kwargs['cambridge']:
                stack.enter_context(self._en_transcription_context_manager())

            affected = self._generate(force, SourceWord)

        return affected


class SplinterTable(Table):
    _NAME = 'splinter'
    _FIELDS = ('nu_graphic', 'first_language', 'survey_language', 'image_id', 'type_of_splinter',
               'sw1_splinter', 'sw2_splinter', 'sw3_splinter', 'sw4_splinter',
               'sw1_splinter_len', 'sw2_splinter_len', 'sw3_splinter_len', 'sw4_splinter_len',
               'G_sw1_splinter', 'G_sw1_splinter__ignore', 'G_sw2_splinter', 'G_sw2_splinter__ignore',
               'G_sw3_splinter', 'G_sw3_splinter__ignore', 'G_sw4_splinter', 'G_sw4_splinter__ignore',
               'G_sw1_splinter_len', 'G_sw1_splinter_len__ignore', 'G_sw2_splinter_len', 'G_sw2_splinter_len__ignore',
               'G_sw3_splinter_len', 'G_sw3_splinter_len__ignore', 'G_sw4_splinter_len', 'G_sw4_splinter_len__ignore')
    _PRIMARY = 5

    _INTEGRITY_SELECT = "SELECT {} FROM naming_unit, ({}) T".format(
        ', '.join(_FIELDS[:_PRIMARY]),
        ' UNION '.join(
            'SELECT \'{}\' type_of_splinter'.format(t) for t in (
                'graphic strict',
                'graphic modified',
                'phonetic strict',
                'phonetic modified'
            )
        )
    )
    # SELECT nu_graphic, first_language, survey_language, image_id, type_of_splinter FROM naming_unit,
    #   (SELECT 'graphic strict' type_of_splinter
    # UNION SELECT 'graphic modified' type_of_splinter
    # UNION SELECT 'phonetic strict' type_of_splinter
    # UNION SELECT 'phonetic modified' type_of_splinter) T;

    _GENERATE_SELECT_ALL = """
SELECT
  SPL.*, NU.nu_phonetic, NU.sw1_graphic, NU.sw2_graphic, NU.sw3_graphic, NU.sw4_graphic,
  SW1.sw_phonetic sw1_phonetic, SW2.sw_phonetic sw2_phonetic, SW3.sw_phonetic sw3_phonetic, SW4.sw_phonetic sw4_phonetic
FROM splinter SPL
  LEFT JOIN naming_unit NU
    ON NU.nu_graphic=SPL.nu_graphic AND NU.first_language=SPL.first_language AND
       NU.survey_language=SPL.survey_language AND NU.image_id=SPL.image_id
  LEFT JOIN source_word SW1
    ON SW1.sw_graphic=NU.sw1_graphic AND
       SW1.first_language=SPL.first_language AND SW1.survey_language=SPL.survey_language
  LEFT JOIN source_word SW2
    ON SW2.sw_graphic=NU.sw2_graphic AND
       SW2.first_language=SPL.first_language AND SW2.survey_language=SPL.survey_language
  LEFT JOIN source_word SW3
    ON SW3.sw_graphic=NU.sw3_graphic AND
       SW3.first_language=SPL.first_language AND SW3.survey_language=SPL.survey_language
  LEFT JOIN source_word SW4
    ON SW4.sw_graphic=NU.sw4_graphic AND
       SW4.first_language=SPL.first_language AND SW4.survey_language=SPL.survey_language
"""

    _GENERATE_SELECT_NEEDED = _GENERATE_SELECT_ALL + """
WHERE
  SPL.type_of_splinter LIKE 'graphic %' AND (
    SPL.G_sw1_splinter IS NULL AND SW1.sw_graphic IS NOT NULL
    OR SPL.G_sw2_splinter IS NULL AND SW2.sw_graphic IS NOT NULL
    OR SPL.G_sw3_splinter IS NULL AND SW3.sw_graphic IS NOT NULL
    OR SPL.G_sw4_splinter IS NULL AND SW4.sw_graphic IS NOT NULL
  ) OR SPL.type_of_splinter LIKE 'phonetic %' AND (
    SPL.G_sw1_splinter IS NULL AND SW1.sw_phonetic IS NOT NULL
    OR SPL.G_sw2_splinter IS NULL AND SW2.sw_phonetic IS NOT NULL
    OR SPL.G_sw3_splinter IS NULL AND SW3.sw_phonetic IS NOT NULL
    OR SPL.G_sw4_splinter IS NULL AND SW4.sw_phonetic IS NOT NULL
  )
"""

    # noinspection PyUnusedLocal
    def generate(self, force, **kwargs) -> int:
        return self._generate(force, Splinter)


class SplinterView(EditableTableLike):

    _NAME = 'splinter_view'

    # nu_graphic, first_language, survey_language, image_id
    _PRIMARY = 4

    _EXCLUDE_EDITABLE = {
        'gs_name',
        'gm_name',
        'ps_name',
        'pm_name',
    }

    # ktore stlpce z naming unit tabulky chceme
    __NU_FIELDS = (
        'nu_graphic', 'first_language', 'survey_language', 'image_id',
        'wf_process', 'wfp_specification',
        'nu_word_class', 'nu_phonetic',
        'nu_syllabic',
        'nu_graphic_len',
        'nu_phonetic_len',
        'nu_syllabic_len',
        'n_of_overlapping_letters',
        'n_of_overlapping_phones',
        'lexsh_main', 'lexsh_sm', 'lexsh_whatm',
        'split_point_1', 'split_point_2', 'split_point_3',
    )

    # ktore stlpce z image tabulky chceme
    __IMG_FIELDS = (
        'sub_sem_cat', 'dom_sem_cat', 'sub_name', 'dom_name',
        'sub_number', 'dom_number', 'half_number', 'sub_sub'
    )

    # ktore stlpce zo source_word maju byt pri kazdom zdrojovom slove
    __SW_FIELDS = (
        'sw_graphic', 'source_language',
        'sw_phonetic', 'sw_word_class', 'sw_syllabic',
        'sw_graphic_len', 'sw_phonetic_len', 'sw_syllabic_len',
    )

    # ktore stlpce zo splinter chceme pre kazdy splinter
    __SPL_FIELDS = (
        'type_of_splinter',
        'sw1_splinter', 'sw2_splinter', 'sw3_splinter', 'sw4_splinter',
        'sw1_splinter_len', 'sw2_splinter_len', 'sw3_splinter_len', 'sw4_splinter_len',
        'G_sw1_splinter', 'G_sw1_splinter__ignore', 'G_sw2_splinter', 'G_sw2_splinter__ignore',
        'G_sw3_splinter', 'G_sw3_splinter__ignore', 'G_sw4_splinter', 'G_sw4_splinter__ignore',
        'G_sw1_splinter_len', 'G_sw1_splinter_len__ignore', 'G_sw2_splinter_len', 'G_sw2_splinter_len__ignore',
        'G_sw3_splinter_len', 'G_sw3_splinter_len__ignore', 'G_sw4_splinter_len', 'G_sw4_splinter_len__ignore'
    )

    # typy splintrov
    __SPL_TYPES = ('gs', 'gm', 'ps', 'pm')

    def __init__(self, wb, conn, as_affected: bool = False):

        self.__field_manager = SplinterViewFieldManager(
            self.__NU_FIELDS, self.__IMG_FIELDS, self.__SW_FIELDS, self.__SPL_FIELDS, self.__SPL_TYPES
        )

        self._FIELDS = self.__field_manager.flat_fields

        self._EXCLUDE_EDITABLE.update(self.__field_manager.static_fields)

        self._EXPORT_SELECT = self.__export_select

        super().__init__(wb, conn, as_affected)  # vyzaduje _FIELDS

        self.__spl_table = SplinterTable(wb, conn, as_affected)

    @property
    def __export_select(self):
        return """SELECT {} FROM naming_unit NU

  LEFT JOIN image I
    ON NU.image_id = I.image_id

  LEFT JOIN source_word SW1
    ON NU.first_language = SW1.first_language
       AND NU.survey_language = SW1.survey_language
       AND NU.sw1_graphic = SW1.sw_graphic
  LEFT JOIN source_word SW2
    ON NU.first_language = SW2.first_language
       AND NU.survey_language = SW2.survey_language
       AND NU.sw2_graphic = SW2.sw_graphic
  LEFT JOIN source_word SW3
    ON NU.first_language = SW3.first_language
       AND NU.survey_language = SW3.survey_language
       AND NU.sw3_graphic = SW3.sw_graphic
  LEFT JOIN source_word SW4
    ON NU.first_language = SW4.first_language
       AND NU.survey_language = SW4.survey_language
       AND NU.sw4_graphic = SW4.sw_graphic

  LEFT JOIN splinter GS
    ON NU.nu_graphic = GS.nu_graphic
      AND NU.first_language = GS.first_language
      AND NU.survey_language = GS.survey_language
      AND NU.image_id = GS.image_id
      AND GS.type_of_splinter = 'graphic strict'

  LEFT JOIN splinter GM
    ON NU.nu_graphic = GM.nu_graphic
      AND NU.first_language = GM.first_language
      AND NU.survey_language = GM.survey_language
      AND NU.image_id = GM.image_id
      AND GM.type_of_splinter = 'graphic modified'

  LEFT JOIN splinter PS
    ON NU.nu_graphic = PS.nu_graphic
      AND NU.first_language = PS.first_language
      AND NU.survey_language = PS.survey_language
      AND NU.image_id = PS.image_id
      AND PS.type_of_splinter = 'phonetic strict'

  LEFT JOIN splinter PM
    ON NU.nu_graphic = PM.nu_graphic
      AND NU.first_language = PM.first_language
      AND NU.survey_language = PM.survey_language
      AND NU.image_id = PM.image_id
      AND PM.type_of_splinter = 'phonetic modified';
""".format(
            ', '.join(self.__field_manager.select_fields)
        )

    def _update(self, data: Dict[str, Any], fields_to_modify: List[str]):

        # nu_graphic, first_language, survey_language, image_id
        base_dict = {f: value for f, value in data.items() if f in self.primary_fields}

        for spl_type in self.__SPL_TYPES:

            to_modify = [
                orig_field
                for orig_field, curr_field in self.__field_manager.original_spl_to_current[spl_type].items()
                if curr_field in fields_to_modify
            ]

            if len(to_modify) > 0:

                data_dict = dict(base_dict)

                # add type_of_splinter, sw1_splinter, ...
                data_dict.update({
                    orig_field: data[curr_field]
                    for orig_field, curr_field in self.__field_manager.original_spl_to_current[spl_type].items()
                })

                # noinspection PyProtectedMember
                self.__spl_table._update(data_dict, to_modify)

    @property
    def integrity_kept(self) -> bool:
        query = """SELECT COUNT(*) FROM naming_unit NU
  LEFT JOIN splinter GS
    ON NU.nu_graphic = GS.nu_graphic
      AND NU.first_language = GS.first_language
      AND NU.survey_language = GS.survey_language
      AND NU.image_id = GS.image_id
      AND GS.type_of_splinter = 'graphic strict'

  LEFT JOIN splinter GM
    ON NU.nu_graphic = GM.nu_graphic
      AND NU.first_language = GM.first_language
      AND NU.survey_language = GM.survey_language
      AND NU.image_id = GM.image_id
      AND GM.type_of_splinter = 'graphic modified'

  LEFT JOIN splinter PS
    ON NU.nu_graphic = PS.nu_graphic
      AND NU.first_language = PS.first_language
      AND NU.survey_language = PS.survey_language
      AND NU.image_id = PS.image_id
      AND PS.type_of_splinter = 'phonetic strict'

  LEFT JOIN splinter PM
    ON NU.nu_graphic = PM.nu_graphic
      AND NU.first_language = PM.first_language
      AND NU.survey_language = PM.survey_language
      AND NU.image_id = PM.image_id
      AND PM.type_of_splinter = 'phonetic modified'

  WHERE GS.type_of_splinter IS NULL OR GM.type_of_splinter IS NULL
    OR PS.type_of_splinter IS NULL OR PM.type_of_splinter IS NULL"""

        return self._execute(query).cursor.fetchone()[0] == 0


class ResponseView(EditableTableLike):

    _NAME = 'response'

    _FIELDS = ('respondent_id', 'image_id', 'nu_original', 'nu_modified')

    _PRIMARY = 4

    _EXPORT_SELECT = """
SELECT O.respondent_id, O.image_id, nu_original, nu_modified
FROM response_original O
LEFT JOIN response_modified M ON O.respondent_id=M.respondent_id AND O.image_id=M.image_id
"""

    @property
    def _emphasized_columns(self) -> Iterable[int]:
        return range(3)  # nu_modified je technicky primarny, ale moze sa upravovat

    def _update(self, data: Dict[str, Any], fields_to_modify: List[str]):
        raise NotImplementedError

    @staticmethod
    def __sheet_rows_013(sheet: Worksheet):
        """Vrati generator na (respondent_id: str, image_id: int, nu_modified: str), kde nu_modified je vyplnene."""
        return (
            (str(row[0].value), int(row[1].value), str(row[3].value))
            for row in sheet.iter_rows(min_row=2)
            if row[3].value
        )

    @staticmethod
    def __row_groupby_key(row: List[Cell]) -> (str, int):
        return str(row[0].value), int(row[1].value)

    def __sync_modified(self, sheet: Worksheet, db_keys: Set[Tuple[str, int]]) -> bool:
        """
        Synchronizuje databazovu tabulku `response_modified`.
        :param sheet: harok
        :param db_keys: dvojice (respondent_id, image_id) z databazovej tabulky response_original
        :raise ResponseDuplicatesException: Ak su v SHEETE duplikaty; parameter je zoznam duplikatov.
        :raise ResponseTypeError: Ak sa pri zoradzovani riadkov v SHEETE vyskytne TypeError.
        :return: pridali alebo zmazali sa riadky v DB?
        """
        modified = False
        # (respondent_id, image_id, nu_modified)
        db_set = set(self._execute("SELECT respondent_id, image_id, nu_modified FROM response_modified").cursor)

        # (respondent_id, image_id, nu_modified)
        try:
            # filtrovanie je dolezite kvoli tomu, aby sa do databazovej tabulky response_modified nanahral zaznam
            # s dvojicou (respondent_id, image_id), ktora nie je v databazovej tabulke respondent_original
            sheet_rows = sorted(filter(
                lambda row: row[:2] in db_keys,
                self.__sheet_rows_013(sheet)
            ))
        except TypeError as e:
            raise ResponseTypeError(*e.args)

        duplicates = [key for key, group in groupby(sheet_rows) if len(list(group)) > 1]
        if len(duplicates) > 0:
            raise ResponseDuplicatesException(duplicates)

        # (respondent_id, image_id, nu_modified)
        sheet_set = set(sheet_rows)

        db_items_only = db_set - sheet_set
        sheet_items_only = sheet_set - db_set

        # co nie je v SHEET -> zmazat v DB
        if len(db_items_only) > 0:
            res = self._executemany(
                "DELETE FROM response_modified WHERE respondent_id = %s AND image_id = %s AND nu_modified = %s",
                list(db_items_only)  # v API je sequence of sequences; set nie je sequence
            )
            if res.result > 0:
                modified = True

        # co nie je v DB -> pridat do DB
        if len(sheet_items_only) > 0:
            res = self._executemany(
                "INSERT INTO response_modified (respondent_id, image_id, nu_modified) VALUES (%s, %s, %s)",
                list(sheet_items_only)
            )
            if res.result > 0:
                modified = True

        return modified

    def __sync_original(self, sheet: Worksheet) -> (bool, Set[Tuple[str, int]]):
        """
        Synchronizuje databazovu tabulku response_original, ktora je read-only.
        :param sheet: harok
        :return: udiala sa nejaka zmena?
        """

        modified = False

        # (respondent_id, image_id) -> nu_original
        db_dict = {dbvals[:2]: dbvals[2] for dbvals in self._execute(
            "SELECT respondent_id, image_id, nu_original FROM response_original"
        ).cursor}

        # (respondent_id, image_id) -> [List[Cell]]
        try:
            sheet_dict = {
                k: list(g) for k, g in groupby(
                    sorted(sheet.iter_rows(min_row=2), key=self.__row_groupby_key),
                    key=self.__row_groupby_key
                )
            }
        except TypeError as e:
            raise ResponseTypeError(*e.args)

        # porovnat mnoziny klucov
        keys_db_only = set(db_dict.keys()) - set(sheet_dict.keys())
        keys_sheet_only = set(sheet_dict.keys()) - set(db_dict.keys())
        keys_both = set(db_dict.keys()) & set(sheet_dict.keys())

        # co su iba v DB, pridat zlto
        if len(keys_db_only) > 0:
            modified = True
            starting_row = sheet.max_row + 1
            for key in keys_db_only:
                sheet.append([*key, db_dict[key], None])
            for row in sheet.iter_rows(min_row=starting_row):
                for cell in row:
                    cell.fill = self._YELLOWFILL

        # co su iba v SHEET, ofarbit cerveno
        if len(keys_sheet_only) > 0:
            for key in keys_sheet_only:
                for row in sheet_dict[key]:
                    for cell in row:
                        if not self._same_fill(cell, self._REDFILL):
                            cell.fill = self._REDFILL
                            modified = True

        # co su aj aj, skontrolovat ci sa neurobil preklep v nu_original
        if len(keys_both) > 0:
            for key in keys_both:
                nu = db_dict[key]
                for row in sheet_dict[key]:
                    cell = row[2]  # nu_original
                    if cell.value != nu:
                        cell.value = nu
                        cell.fill = self._YELLOWFILL
                        modified = True

        return modified, set(db_dict.keys())

    def sync(self, unhighlight: bool = False) -> bool:

        sheet = self._wb[self._NAME]

        # odstranit zvyraznenie minulych zmien
        if unhighlight:
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    if self._same_fill(cell, self._YELLOWFILL):
                        cell.fill = self._NOFILL

        modified = False

        res, db_keys = self.__sync_original(sheet)

        if res:
            modified = True

        if self.__sync_modified(sheet, db_keys):
            modified = True

        return modified

    @property
    def integrity_kept(self) -> bool:
        resultrow = self._execute("""
SELECT
  (
    SELECT COUNT(*) FROM respondent R
    LEFT JOIN response_original O ON R.respondent_id = O.respondent_id
    WHERE O.respondent_id IS NULL
  ) + (
    SELECT COUNT(*) FROM respondent R
    RIGHT JOIN response_original O ON R.respondent_id = O.respondent_id
    WHERE R.respondent_id IS NULL
  ) = 0 AS integrity_kept
""").cursor.fetchone()
        return resultrow[0] == 1
