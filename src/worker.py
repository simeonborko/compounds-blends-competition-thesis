import warnings
from datetime import datetime
from genericpath import isfile
from shutil import copyfile
from tkinter import NORMAL, messagebox, DISABLED

from openpyxl import Workbook, load_workbook

from src import configuration
from src.syncmanager import SyncManager
from src.tools import Connection
from src.tools.exception import ResponseDuplicatesException, ResponseTypeError


def __msg(syncmanager: SyncManager, msgs=None) -> str:
    if msgs is None:
        msgs = []

    clss = syncmanager.modified
    if len(clss) > 0:
        msgs.append('Zmenené:')
        msgs.append('\n'.join('- ' + cls.name() for cls in clss))

    clss = syncmanager.stayed
    if len(clss) > 0:
        msgs.append('Bez zmeny:')
        msgs.append('\n'.join('- ' + cls.name() for cls in clss))

    clss = syncmanager.affected
    if len(clss) > 0:
        msgs.append('Zmenené v dôsledku:')
        msgs.append('\n'.join('- ' + cls.name() for cls in clss))

    return '\n\n'.join(msgs)


class Disabler:
    def __init__(self, buttons, checkers):
        self.__buttons = buttons
        self.__checkers = checkers

    def __enter__(self):
        for button in self.__buttons:
            button.configure(state=DISABLED)
        for checker in self.__checkers:
            checker.configure(state=DISABLED)

    def __exit__(self, exc_type, exc_val, exc_tb):
        # ak je vynimka, otvorit okienko s chybou
        if exc_type is not None:
            messagebox.showerror('Chyba', exc_type.__name__)

        for button in self.__buttons:
            button.configure(state=NORMAL)
        for checker in self.__checkers:
            checker.configure(state=NORMAL)


def backup():
    copyfile(
        configuration.XLSX_FILE,
        configuration.BACKUP_FILE_FMT.format(datetime.now().strftime("%Y-%m-%d--%H-%M-%S"))
    )


def export(clss, widgets):
    with Disabler(**widgets):
        existed = isfile(configuration.XLSX_FILE)

        if not existed:
            wb = Workbook()
            wb.remove(wb['Sheet'])
        else:
            wb = load_workbook(configuration.XLSX_FILE)

        created = []
        notcreated = []
        with Connection() as conn:
            for cls in clss:
                (created if cls(wb, conn).create_sheet() else notcreated).append(cls)

        wb.save(configuration.XLSX_FILE)

        if not existed:
            messagebox.showinfo(
                'Exportovať',
                'Dokument vytvorený:\n\n{}'.format(configuration.XLSX_FILE)
            )
        elif len(created) > 0:
            messagebox.showinfo(
                'Exportovať',
                'Boli pridané tieto hárky:\n\n{}'.format(
                    '\n'.join('- ' + cls.name() for cls in created)
                )
            )
        else:
            messagebox.showwarning(
                'Exportovať',
                'Dokument bol už vytvorený a žiadne hárky neboli pridané'
            )


def sync(clss, unhighlight: bool, widgets):
    with Disabler(**widgets):
        # kontrola ci subor existuje
        if not isfile(configuration.XLSX_FILE):
            messagebox.showerror('Chyba', 'Súbor {} neexistuje'.format(configuration.XLSX_FILE))
            return
        backup()

        if len(clss) == 0:
            messagebox.showwarning(
                'Synchronizácia',
                'Neboli vybrané žiadne tabuľky'
            )
            return

        wb = load_workbook(configuration.XLSX_FILE)
        with Connection() as conn:
            syncmanager = SyncManager(clss, wb, conn)
            syncmanager.sync(unhighlight)
            conn.commit()

        wb.save(configuration.XLSX_FILE)

        messagebox.showinfo('Synchronizácia', __msg(syncmanager))


def generate(clss, unhighlight: bool, force, corpus, bnc_corpus, cambridge, splinter_derived, widgets):
    with Disabler(**widgets):
        if not isfile(configuration.XLSX_FILE):
            messagebox.showerror('Chyba', 'Súbor {} neexistuje'.format(configuration.XLSX_FILE))
            return
        backup()

        if len(clss) == 0:
            messagebox.showwarning(
                'Automatizované vyplnenie',
                'Neboli vybrané žiadne tabuľky'
            )
            return

        wb = load_workbook(configuration.XLSX_FILE)
        with Connection() as conn:
            syncmanager = SyncManager(clss, wb, conn)
            syncmanager.generate(unhighlight, force=force, bnc_corpus=bnc_corpus, corpus=corpus, cambridge=cambridge, splinter_derived=splinter_derived)
            conn.commit()

        wb.save(configuration.XLSX_FILE)

        messagebox.showinfo('Automatizované vyplnenie', __msg(syncmanager))


def integrity(clss, unhighlight: bool, widgets):
    with Disabler(**widgets):
        if not isfile(configuration.XLSX_FILE):
            messagebox.showerror('Chyba', 'Súbor {} neexistuje'.format(configuration.XLSX_FILE))
            return
        backup()

        if len(clss) == 0:
            messagebox.showwarning(
                'Kontrola súdržnosti',
                'Neboli vybrané žiadne tabuľky'
            )
            return

        warnings.filterwarnings('ignore', r'.*Duplicate entry')

        wb = load_workbook(configuration.XLSX_FILE)
        with Connection() as conn:
            syncmanager = SyncManager(clss, wb, conn)
            result = syncmanager.integrity(unhighlight)
            conn.commit()

        warnings.resetwarnings()

        wb.save(configuration.XLSX_FILE)

        msg = __msg(syncmanager, ['\n'.join(
            '- {} (pridané: {}, odobrané: {})'.format(name, *args) for name, args in result.items()
        )])
        messagebox.showinfo('Kontrola súdržnosti', msg)


def splinterview(cls, unhighlight: bool, widgets):
    with Disabler(**widgets):
        if not isfile(configuration.XLSX_FILE):
            messagebox.showerror('Chyba', 'Súbor {} neexistuje'.format(configuration.XLSX_FILE))
            return
        backup()

        title = 'Splinter View'

        wb = load_workbook(configuration.XLSX_FILE)
        with Connection() as conn:
            obj = cls(wb, conn)
            if not obj.integrity_kept:
                messagebox.showerror(title, 'Chyba. Skontroluj súdržnosť pre tabuľku splinter.')
                return

            # ak neexistuje, tak vytvorime
            if cls.name() not in wb.sheetnames:
                if cls(wb, conn).create_sheet():
                    wb.save(configuration.XLSX_FILE)
                    messagebox.showinfo(title, 'Hárok bol vytvorený')
                else:
                    messagebox.showerror(title, 'Hárok sa nepodarilo vytvoriť')

            else:
                syncmanager = SyncManager([cls], wb, conn)
                syncmanager.sync(unhighlight)
                conn.commit()
                wb.save(configuration.XLSX_FILE)
                messagebox.showinfo(title, __msg(syncmanager))


def overview(cls, unhighlight: bool, widgets):
    with Disabler(**widgets):
        if not isfile(configuration.XLSX_FILE):
            messagebox.showerror('Chyba', 'Súbor {} neexistuje'.format(configuration.XLSX_FILE))
            return
        backup()

        title = 'Overview'

        wb = load_workbook(configuration.XLSX_FILE)

        # ak neexistuje, tak vytvorime
        if cls.name() not in wb.sheetnames:
            with Connection() as conn:
                result = cls(wb, conn).create_sheet()
            if result:
                wb.save(configuration.XLSX_FILE)
                messagebox.showinfo(title, 'Hárok bol vytvorený')
            else:
                messagebox.showerror(title, 'Hárok sa nepodarilo vytvoriť')

        # ak existuje, tak synchronizujeme
        else:
            with Connection() as conn:
                syncmanager = SyncManager([cls], wb, conn)
                syncmanager.sync(unhighlight)
                conn.commit()
            wb.save(configuration.XLSX_FILE)

            messagebox.showinfo(title, __msg(syncmanager))


def responseview(cls, unhighlight: bool, widgets):
    with Disabler(**widgets):
        if not isfile(configuration.XLSX_FILE):
            messagebox.showerror('Chyba', 'Súbor {} neexistuje'.format(configuration.XLSX_FILE))
            return
        backup()
        title = 'Response'

        wb = load_workbook(configuration.XLSX_FILE)
        with Connection() as conn:
            obj = cls(wb, conn)
            if not obj.integrity_kept:
                messagebox.showerror(title, 'Chyba. Volaj Simeona.')
                return

            # ak neexistuje, vytvorime
            if cls.name() not in wb.sheetnames:
                if cls(wb, conn).create_sheet():
                    wb.save(configuration.XLSX_FILE)
                    messagebox.showinfo(title, 'Hárok bol vytvorený')
                else:
                    messagebox.showerror(title, 'Hárok sa nepodarilo vytvoriť')

            else:
                try:
                    syncmanager = SyncManager([cls], wb, conn)
                    syncmanager.sync(unhighlight)
                    conn.commit()
                    wb.save(configuration.XLSX_FILE)
                    messagebox.showinfo(title, __msg(syncmanager))
                except ResponseDuplicatesException as e:
                    messagebox.showerror(title, 'Chyba - našli sa duplikáty:\n\n{}'.format(
                        '\n'.join('- ' + str(r) for r in e.args[0])
                    ))
                except ResponseTypeError as e:
                    messagebox.showerror(
                        title,
                        'Chyba: {}\n\nPravdepodobne má hárok v stĺpci respondent_id číselnú hodnotu, čo je chyba.'
                        .format(str(e))
                    )
